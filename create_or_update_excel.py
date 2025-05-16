API_KEY = "ak-YOUR_API_KEY_GOES_HERE"

import requests
import json
import pkg_resources
import csv
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter


def get_content_type(url):
    try:
        # Send a GET request with a small timeout and limit the response content size
        response = requests.get(url, stream=True, timeout=5)
        
        # Check if Content-Type header exists
        content_type = response.headers.get('Content-Type', '').lower()
        return content_type
        # Check if Content-Type indicates a file type
    except requests.RequestException as e:
        return f"Error accessing {url}: {e}"

def url_creation(file_bytes, file_name="output.xlsx"):
    url = "https://prodaus.api.airia.ai/v1/upload"
    headers = {
        "X-API-Key": API_KEY
    }

    files = {'file': (file_name, file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

    response = requests.post(url, headers=headers, files=files)
    if response.status_code == 200:
        dictionary = json.loads(response.text)
        return dictionary['imageUrl']
    else:
        return f"Failed to upload the file. Status code: {response.status_code}, Response: {response.text}"

def create_excel_from_csv(csv_text, output_excel_file, max_col_width=30):
    """
    Create an Excel file from scratch and initialize the page with a CSV given.
    Automatically adjusts column width, applies wrap text, and highlights the header row.
    :param csv_text: CSV text (2 lines, separated by newline) as input
    :param output_excel_file: Path to the output Excel file
    :param max_col_width: Maximum width for any column
    """
    # Parse CSV text
    csv_lines = csv_text.strip().split("\n")
    csv_reader = csv.reader(csv_lines)
    
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write CSV data to the Excel sheet and adjust column widths
    for row_index, row_data in enumerate(csv_reader, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=cell_value)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Apply wrap text
            
            # Highlight the header row
            if row_index == 1:
                cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Adjust column widths
    for col_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, max_col_width)
        sheet.column_dimensions[get_column_letter(col_index)].width = adjusted_width

    # Save the Excel file
    workbook.save(output_excel_file)
    print(f"Excel file created and initialized with CSV data: {output_excel_file}")

def append_csv_to_excel(csv_text, excel_url, updated_excel_file, max_col_width=50):
    """
    Open an existing Excel file from a URL and append the 2nd line from the CSV to the table.
    Automatically adjusts column width and applies wrap text.
    :param csv_text: CSV text (2 lines, separated by newline) as input
    :param excel_url: URL to the existing Excel file
    :param max_col_width: Maximum width for any column
    """
    # Parse the second line of the CSV text
    csv_lines = csv_text.strip().split("\n")
    if len(csv_lines) < 2:
        raise ValueError("CSV text must have at least two lines.")

    second_line = csv_lines[1]
    csv_reader = csv.reader([second_line])
    row_to_append = next(csv_reader)

    # Download the existing Excel file from the URL
    response = requests.get(excel_url)
    if response.status_code != 200:
        raise ValueError("Failed to download the Excel file.")

    excel_data = BytesIO(response.content)

    # Load the Excel file from the downloaded content
    workbook = openpyxl.load_workbook(excel_data)
    sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Append the data and apply wrap text
    for col_index, cell_value in enumerate(row_to_append, start=1):
        cell = sheet.cell(row=next_row, column=col_index, value=cell_value)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Apply wrap text

    # Adjust column widths
    for col_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, max_col_width)
        sheet.column_dimensions[get_column_letter(col_index)].width = adjusted_width

    # Save the updated Excel file locally (optional, as re-uploading to a server may be required)
    workbook.save(updated_excel_file)
    print(f"Appended data to Excel file and saved as: {updated_excel_file}")

# Example usage
if __name__ == "__main__":
    if len(client_data["images"]) == 0:
        raise "No Files Attached"

    output_excel_file = "output.xlsx"
    excel_url = None
    content = json.loads(input)
    excel_url = content["url"]
    csv_content = content["csv"]
    output = ""
    
    for url in client_data["images"]:
        if get_content_type(url) == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            excel_url = url
            break

    if excel_url is None or len(excel_url) == 0:
        create_excel_from_csv(csv_content, output_excel_file)
    else:
        append_csv_to_excel(csv_content, excel_url, output_excel_file)

    with open(output_excel_file, "rb") as file:
        file_bytes = file.read()
    output = url_creation(file_bytes, file_name=output_excel_file)
    
    setValue("images", [""])
